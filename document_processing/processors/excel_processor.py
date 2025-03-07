# processors/excel_processor.py
import pandas as pd
import openpyxl
from .base_processor import BaseProcessor
from .exceptions import DocumentProcessingError
from pathlib import Path


class ExcelProcessor(BaseProcessor):
    @classmethod
    def extract_text(cls, file_path: str) -> str:
        """流式读取大文件，优化内存使用"""
        try:
            output = []
            # 使用openpyxl优化大文件读取
            wb = openpyxl.load_workbook(
                filename=file_path,
                read_only=True,
                data_only=True
            )

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row == 0:  # 空Sheet跳过
                    continue

                # 转换为Markdown表格
                table_data = []
                for row in ws.iter_rows(values_only=True):
                    table_data.append([str(cell) if cell is not None else "" for cell in row])

                if table_data:
                    output.append(f"=== Sheet '{sheet_name}' ===")
                    output.append(
                        cls._convert_to_markdown(table_data)
                    )

            return "\n\n".join(output)

        except Exception as e:
            cls.safe_logging(file_path, e)
            raise DocumentProcessingError(f"Excel处理失败: {str(e)}")

    @staticmethod
    def _convert_to_markdown(data: list) -> str:
        """高效生成Markdown表格（避免pandas内存开销）"""
        headers = data[0]
        sep = ["---"] * len(headers)
        rows = [headers, sep] + data[1:]
        return "\n".join("| " + " | ".join(row) + " |" for row in rows)